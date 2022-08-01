import zipfile
from zipfile import ZipFile
import os
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from files_scripts.read_pdf import reader
import csv

# Create .zip file
zip_object = ZipFile('sample.zip', 'w')
zip_object.write('docs-pytest-org-en-latest.pdf')
zip_object.write('file_example_XLSX_50.xlsx')
zip_object.write('report.csv')
zip_object.close()

# Move zip to resources folder
source_files = '/Users/vladkim/Desktop/qaguru_work_with_files/sample.zip'
destination_folder = '/Users/vladkim/Desktop/qaguru_work_with_files/resources/sample.zip'
os.rename(source_files, destination_folder)

def test_show_list_of_files_in_zip():
    zip_ = ZipFile('resources/sample.zip')
    print(zip_.namelist())

def test_unzip_files():
    open_zip = zipfile.ZipFile('resources/sample.zip')
    open_zip.extractall('unzipped/')
    open_zip.close()

def test_read_pdf_len():
    reader = PdfReader('unzipped/docs-pytest-org-en-latest.pdf')
    number_of_pages = len(reader.pages)
    print(number_of_pages)

def test_read_pdf():
    page = reader.pages[2]
    text = page.extract_text()
    print(text)
    assert 'styleset' in text

def test_read_xlsx():
    workbook = load_workbook('unzipped/file_example_XLSX_50.xlsx')
    sheet = workbook.active
    print(sheet.cell(row=3, column=4).value)

def test_read_csv():
    with open('unzipped/report.csv', 'r') as f:
        reader = csv.reader(f)
        for row in reader:
            print(row)








